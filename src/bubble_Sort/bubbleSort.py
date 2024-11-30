import os
import time  # Para medir o tempo de execução
import platform  # Para pegar informações do sistema e da linguagem
import psutil  # Para medir o uso de memória
import statistics  # Para calcular média e mediana
from openpyxl import Workbook  # Para criar a planilha Excel
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def bubble_sort(numbers):
    n = len(numbers)
    for i in range(n):
        for j in range(0, n - i - 1):
            
            if numbers[j] > numbers[j + 1]:
                numbers[j], numbers[j + 1] = numbers[j + 1], numbers[j]

#ler o arquivo
def read_numbers_from_file(filename):
    with open(filename, 'r') as file:
        
        return [int(line.strip()) for line in file]

def write_numbers_to_file(numbers, filename):
    with open(filename, 'w') as file:
        for number in numbers:
            file.write(f"{number}\n")  

# Salva os resultados na planilha Excel
def save_to_excel(execution_times, memory_usages, avg_time, median_time, avg_memory, median_memory):
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados Bubble Sort"

    headers = ["Execução", "Tempo (ms)", "Memória (KB)"]
    ws.append(headers)

    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"].font = header_font
        ws[f"{col_letter}1"].alignment = Alignment(horizontal="center")

    for i in range(len(execution_times)):
        ws.append([i + 1, execution_times[i], memory_usages[i]])

    ws.append(["", "", ""])  
    ws.append(["Estatísticas", "Tempo (ms)", "Memória (KB)"])
    ws.append(["Média", avg_time, avg_memory])
    ws.append(["Mediana", median_time, median_memory])

    for col in range(2, 4):
        for row in range(2, len(execution_times) + 5):
            ws[f"{get_column_letter(col)}{row}"].alignment = Alignment(horizontal="right")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except TypeError:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save("resultadosPYTHON_BUBBLE.xlsx")
    print("Resultados salvos em resultadosPYTHON_BUBBLE.xlsx.")

def main():
    input_file = "arqTEST03.txt" 
    output_file = "PYarq-saida.txt"  

    print("---====---" * 5)
    print("Ordenação por Bubble Sort")
    print("Linguagem: Python")
    print("Versão:", platform.python_version())

    print("\nInformações do sistema:")
    print("Sistema:", platform.system())
    print("Processador:", platform.processor())
    print("Arquitetura:", platform.architecture()[0])
    print("Memória RAM total:", psutil.virtual_memory().total // 1024, "KB")
    print("Memória Livre:", psutil.virtual_memory().available // 1024, "KB")
    print("Número de CPUs:", psutil.cpu_count(logical=False))
    print("Hostname:", platform.node())
    print("Usuário Atual:", os.getlogin())

    print("\nLendo os números do arquivo...\n")
    numbers = read_numbers_from_file(input_file)
    print("Números lidos:", numbers)

    execution_times = []
    memory_usages = []

    print("\nExecutando Bubble Sort 10 vezes...")
    for i in range(10):

        start_time = time.time() 
        start_memory = psutil.Process().memory_info().rss // 1024  # Marca a memória usada no início

        
        bubble_sort(numbers)

        end_time = time.time()
        end_memory = psutil.Process().memory_info().rss // 1024

        elapsed_time_ms = (end_time - start_time) * 1000  
        memory_used_kb = end_memory - start_memory  

        execution_times.append(elapsed_time_ms)
        memory_usages.append(memory_used_kb)

        print(f"Execução {i + 1}: Tempo = {elapsed_time_ms:.2f} ms, Memória = {memory_used_kb} KB")

    avg_time = statistics.mean(execution_times)
    median_time = statistics.median(execution_times)
    avg_memory = statistics.mean(memory_usages)
    median_memory = statistics.median(memory_usages)

    print("\nSalvando os números ordenados no arquivo...")
    write_numbers_to_file(numbers, output_file)
    print(f"Números ordenados salvos em {output_file}.\n")

    print("---====---" * 5)
    print("Resultados Finais:")
    print(f"Tempo - Média: {avg_time:.2f} ms, Mediana: {median_time:.2f} ms")
    print(f"Memória - Média: {avg_memory} KB, Mediana: {median_memory} KB")
    print("---====---" * 5)
    print("\n\n")

    save_to_excel(execution_times, memory_usages, avg_time, median_time, avg_memory, median_memory)

if __name__ == "__main__":
    main()
